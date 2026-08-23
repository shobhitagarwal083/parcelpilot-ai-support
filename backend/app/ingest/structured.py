"""Workbook to SQLite, with real types and IST-aware timestamps.

The workbook stores every timestamp as a naive `YYYY-MM-DD HH:MM` string. They
are localised to Asia/Kolkata here, once, at the boundary — so that no downstream
module ever has to decide what timezone a bare string is in.

`historical_resolution` is carried through unchanged. It is tier-4 evidence and
two of its values are wrong on purpose; the repository layer is responsible for
never presenting it as authority.
"""

from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

from app import config
from app.ingest.documents import IngestError

WORKBOOK_NAME = "ParcelPilot_Assessment_Data.xlsx"

EXPECTED_COLUMNS = {
    "accounts": [
        "account_id",
        "account_name",
        "plan",
        "status",
        "csm",
        "contract_file",
        "premium_support",
        "notes",
    ],
    "orders": [
        "order_id",
        "account_id",
        "carrier",
        "status",
        "booked_at",
        "pickup_window_start",
        "pickup_window_end",
        "pickup_actual_at",
        "shipment_fee_inr",
        "carrier_fault",
        "customer_fault",
        "cancellation_requested_at",
        "notes",
    ],
    "tickets": [
        "ticket_id",
        "account_id",
        "created_at",
        "status",
        "subject",
        "description",
        "channel",
        "assigned_to",
        "last_customer_message_at",
        "historical_resolution",
    ],
}

TIMESTAMP_COLUMNS = {
    "booked_at",
    "pickup_window_start",
    "pickup_window_end",
    "pickup_actual_at",
    "cancellation_requested_at",
    "created_at",
    "last_customer_message_at",
}

SCHEMA = """
CREATE TABLE accounts (
    account_id      TEXT PRIMARY KEY,
    account_name    TEXT NOT NULL,
    plan            TEXT NOT NULL,
    status          TEXT NOT NULL,
    csm             TEXT,
    contract_file   TEXT,
    premium_support INTEGER NOT NULL,
    notes           TEXT
);

CREATE TABLE orders (
    order_id                  TEXT PRIMARY KEY,
    account_id                TEXT NOT NULL REFERENCES accounts(account_id),
    carrier                   TEXT NOT NULL,
    status                    TEXT NOT NULL,
    booked_at                 TEXT NOT NULL,
    pickup_window_start       TEXT,
    pickup_window_end         TEXT,
    pickup_actual_at          TEXT,
    shipment_fee_inr          REAL NOT NULL,
    -- Non-nullable booleans in the source sheet, so FALSE means established
    -- not-at-fault rather than unknown (assumption A8). The distinction decides
    -- whether a credit answer is `ineligible` or `indeterminate`.
    carrier_fault             INTEGER NOT NULL,
    customer_fault            INTEGER NOT NULL,
    cancellation_requested_at TEXT,
    notes                     TEXT
);

CREATE TABLE tickets (
    ticket_id                TEXT PRIMARY KEY,
    account_id               TEXT NOT NULL REFERENCES accounts(account_id),
    created_at               TEXT NOT NULL,
    status                   TEXT NOT NULL,
    subject                  TEXT NOT NULL,
    description              TEXT,
    channel                  TEXT,
    assigned_to              TEXT,
    -- A customer message, NOT a ParcelPilot response. The workbook has no
    -- first_response_at column at all. The SLA clock starts at created_at and
    -- this column must never be substituted for it: doing so silently moves
    -- TKT-501 from breached to within target.
    last_customer_message_at TEXT,
    -- Tier 4. Context only, and known to contain incorrect past guidance.
    historical_resolution    TEXT
);

CREATE INDEX idx_orders_account ON orders(account_id);
CREATE INDEX idx_tickets_account ON tickets(account_id);
"""


def _parse_timestamp(value: object, *, column: str) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        dt = value if value.tzinfo else value.replace(tzinfo=config.TZ)
        return dt.astimezone(config.TZ).isoformat()
    try:
        naive = datetime.strptime(str(value).strip(), "%Y-%m-%d %H:%M")
    except ValueError as exc:
        raise IngestError(f"{column}: cannot parse timestamp {value!r}") from exc
    return naive.replace(tzinfo=config.TZ).isoformat()


def _rows(sheet, name: str) -> list[dict]:
    raw = list(sheet.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in raw[0]]
    if header != EXPECTED_COLUMNS[name]:
        raise IngestError(
            f"sheet {name!r} has unexpected columns.\n"
            f"  expected: {EXPECTED_COLUMNS[name]}\n"
            f"  found:    {header}"
        )

    rows = []
    for values in raw[1:]:
        if all(v is None or v == "" for v in values):
            continue
        record = {}
        for column, value in zip(header, values, strict=True):
            if column in TIMESTAMP_COLUMNS:
                record[column] = _parse_timestamp(value, column=f"{name}.{column}")
            elif isinstance(value, bool):
                record[column] = int(value)
            elif value == "":
                record[column] = None
            else:
                record[column] = value
        rows.append(record)
    return rows


def parse_snapshot(sheet) -> datetime:
    """Read the pinned reference time out of the workbook's own README sheet."""
    values = {str(r[0]).strip(): r[1] for r in sheet.iter_rows(values_only=True) if r[0]}
    raw = values.get("Dataset snapshot")
    if raw is None:
        raise IngestError("README sheet has no 'Dataset snapshot' row")
    if isinstance(raw, datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=config.TZ)

    stamp, _, zone = str(raw).strip().rpartition(" ")
    return datetime.strptime(stamp, "%Y-%m-%d %H:%M").replace(tzinfo=ZoneInfo(zone))


def ingest_structured(source_dir: Path | None = None, db_path: Path | None = None) -> dict:
    source_dir = source_dir or config.SOURCE_DIR
    db_path = db_path or config.DB_PATH
    workbook_path = source_dir / WORKBOOK_NAME
    if not workbook_path.exists():
        raise IngestError(f"missing workbook: {workbook_path}")

    book = openpyxl.load_workbook(workbook_path, data_only=True)

    snapshot = parse_snapshot(book["README"])
    if snapshot != config._PINNED_SNAPSHOT:
        raise IngestError(
            f"the workbook pins the snapshot at {snapshot.isoformat()} but config._PINNED_SNAPSHOT "
            f"is {config._PINNED_SNAPSHOT.isoformat()}. Every timing answer in the system depends "
            f"on this value; update config.py deliberately rather than letting them drift."
        )

    tables = {name: _rows(book[name], name) for name in EXPECTED_COLUMNS}

    known_accounts = {row["account_id"] for row in tables["accounts"]}
    for name, key in (("orders", "order_id"), ("tickets", "ticket_id")):
        orphans = [r[key] for r in tables[name] if r["account_id"] not in known_accounts]
        if orphans:
            raise IngestError(f"{name} reference unknown accounts: {', '.join(orphans)}")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    db_path.unlink(missing_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.executescript(SCHEMA)
        for name, rows in tables.items():
            columns = EXPECTED_COLUMNS[name]
            conn.executemany(
                f"INSERT INTO {name} ({','.join(columns)}) VALUES ({','.join('?' * len(columns))})",
                [tuple(row[c] for c in columns) for row in rows],
            )

    config.SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
    config.SNAPSHOT_PATH.write_text(
        json.dumps(
            {"snapshot_at": snapshot.isoformat(), "source": f"{WORKBOOK_NAME}!README"}, indent=2
        )
    )

    counts = {name: len(rows) for name, rows in tables.items()}
    return counts | {"snapshot_at": snapshot.isoformat()}
